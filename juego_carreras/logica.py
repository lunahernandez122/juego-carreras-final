# KEVIN HERNÁNDEZ HENAO & LUNA HERNANDEZ MONTOYA
# Juego de Carreras — POO - Fase 2 (lógica)

from __future__ import annotations
from abc import ABC, abstractmethod
from typing import Optional
import random
import json
import csv

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False


class Item:
    """Ítem consumible que ocupa una ranura del inventario."""

    def __init__(self, nombre: str, tipo: str, efecto: int) -> None:
        self.nombre: str = nombre
        self.tipo: str = tipo
        self.efecto: int = efecto

    def __str__(self) -> str:
        return f"[{self.tipo}] {self.nombre} (+{self.efecto}m)"

    def __repr__(self) -> str:
        return f"Item(nombre={self.nombre!r}, tipo={self.tipo!r}, efecto={self.efecto})"

    def __eq__(self, otro: object) -> bool:
        if not isinstance(otro, Item):
            return NotImplemented
        return self.nombre == otro.nombre and self.tipo == otro.tipo


class Inventario:
    """
    Contenedor de ítems. Composición con Vehiculo.
    Implementa __len__, __iter__ y __contains__.
    """

    def __init__(self, capacidad: int = 3) -> None:
        self.__items: list[Item] = []
        self.__capacidad: int = capacidad

    def agregar(self, item: Item) -> bool:
        if len(self.__items) < self.__capacidad:
            self.__items.append(item)
            return True
        return False

    def usar(self, indice: int) -> Optional[Item]:
        if 0 <= indice < len(self.__items):
            return self.__items.pop(indice)
        return None

    def listar(self) -> list[Item]:
        return list(self.__items)

    def __len__(self) -> int:
        return len(self.__items)

    def __iter__(self):
        return iter(self.__items)

    def __contains__(self, item: object) -> bool:
        return item in self.__items

    def __str__(self) -> str:
        if not self.__items:
            return "Vacío"
        return " | ".join(str(i) for i in self.__items)

    def __repr__(self) -> str:
        return f"Inventario(capacidad={self.__capacidad}, items={self.__items!r})"


class Vehiculo(ABC):
    """Clase abstracta base para todos los vehículos del juego."""

    _contador: int = 0

    def __init__(
        self,
        nombre: str,
        marca: str,
        placa: str,
        color: str,
        velocidad: int,
        gasolina: int,
    ) -> None:
        Vehiculo._contador += 1
        self.__nombre: str = nombre
        self.marca: str = marca
        self.placa: str = placa
        self.color: str = color
        self.velocidad: int = velocidad
        self.gasolina: int = gasolina
        self.posicion: int = 0
        self.turbos: int = 3
        self.turbos_usados: int = 0
        self.es_jugador: bool = False
        self.inventario: Inventario = Inventario()

    @property
    def nombre(self) -> str:
        return self.__nombre

    @abstractmethod
    def acelerar(self) -> None:
        """Avanza el vehículo según su tipo."""
        ...

    @abstractmethod
    def habilidad_especial(self) -> None:
        """Habilidad única de cada tipo de vehículo."""
        ...

    @classmethod
    def total_vehiculos(cls) -> int:
        """Retorna cuántos vehículos se han creado en total."""
        return cls._contador

    @classmethod
    def crear_rival(cls, nombre: str) -> "Vehiculo":
        """Constructor alternativo: crea un rival con valores predeterminados."""
        return cls(nombre, "Marca", "ABC-000", "Negro", 100, 100)

    @staticmethod
    def calcular_avance_turbo(velocidad_base: int, factor: float = 1.5) -> int:
        """Calcula el avance extra que da el turbo según la velocidad base."""
        return int(velocidad_base * factor * 0.15)

    @staticmethod
    def clasificar_posicion(posicion: int, distancia: int) -> str:
        """Clasifica qué tan cerca está un vehículo de la meta."""
        porcentaje = (posicion / distancia) * 100 if distancia > 0 else 0
        if porcentaje >= 80:
            return "¡Casi en meta!"
        elif porcentaje >= 50:
            return "A mitad de camino"
        elif porcentaje >= 20:
            return "Arrancando"
        else:
            return "Al inicio"

    def usar_turbo(self) -> bool:
        """Activa el turbo manualmente. Retorna True si se activó."""
        if self.turbos >= 1:
            avance = Vehiculo.calcular_avance_turbo(self.velocidad)
            self.posicion += avance
            self.turbos -= 1
            self.turbos_usados += 1
            return True
        return False

    def turbo_rival(self, prob: float = 0.20) -> bool:
        """El turbo del rival se activa de forma aleatoria."""
        if self.turbos >= 1 and random.random() < prob:
            avance = Vehiculo.calcular_avance_turbo(self.velocidad)
            self.posicion += avance
            self.turbos -= 1
            self.turbos_usados += 1
            return True
        return False

    def __str__(self) -> str:
        etq = " (TÚ)" if self.es_jugador else ""
        return (f"{self.__nombre}{etq} | {self.__class__.__name__} | "
                f"Pos: {self.posicion}m | Turbos: {self.turbos}")

    def __repr__(self) -> str:
        return (f"{type(self).__name__}(nombre={self.__nombre!r}, "
                f"posicion={self.posicion}, turbos={self.turbos})")

    def __eq__(self, otro: object) -> bool:
        if not isinstance(otro, Vehiculo):
            return NotImplemented
        return self.nombre == otro.nombre

    def __lt__(self, otro: "Vehiculo") -> bool:
        return self.posicion < otro.posicion


class Carro(Vehiculo):
    def __init__(
        self,
        nombre: str,
        marca: str = "Mazda",
        placa: str = "ABC123",
        color: str = "Rojo",
        velocidad: int = 120,
        gasolina: int = 100,
        puertas: int = 4,
    ) -> None:
        super().__init__(nombre, marca, placa, color, velocidad, gasolina)
        self.puertas: int = puertas

    def acelerar(self) -> None:
        self.posicion += 10

    def habilidad_especial(self) -> None:
        self.posicion += 5

    def __repr__(self) -> str:
        return f"Carro(nombre={self.nombre!r}, puertas={self.puertas})"


class Moto(Vehiculo):
    def __init__(
        self,
        nombre: str,
        marca: str = "Yamaha",
        placa: str = "DEF456",
        color: str = "Azul",
        velocidad: int = 150,
        gasolina: int = 80,
        tipo: str = "Deportiva",
    ) -> None:
        super().__init__(nombre, marca, placa, color, velocidad, gasolina)
        self.tipo: str = tipo

    def acelerar(self) -> None:
        self.posicion += 15

    def habilidad_especial(self) -> None:
        self.posicion += 15

    def __repr__(self) -> str:
        return f"Moto(nombre={self.nombre!r}, tipo={self.tipo!r})"


class Camion(Vehiculo):
    def __init__(
        self,
        nombre: str,
        marca: str = "Chevrolet",
        placa: str = "GHI789",
        color: str = "Verde",
        velocidad: int = 90,
        gasolina: int = 150,
        carga: int = 5000,
    ) -> None:
        super().__init__(nombre, marca, placa, color, velocidad, gasolina)
        self.carga: int = carga

    def acelerar(self) -> None:
        self.posicion += 5

    def habilidad_especial(self) -> None:
        self.posicion += 8

    def __repr__(self) -> str:
        return f"Camion(nombre={self.nombre!r}, carga={self.carga}kg)"


class Carrera:
    """Gestiona una carrera individual. Composición: contiene una lista de Vehiculo."""

    PUNTOS_F1: dict[int, int] = {
        1: 25, 2: 18, 3: 15, 4: 12, 5: 10,
        6: 8,  7: 6,  8: 4,  9: 2,  10: 1,
    }

    DIFICULTADES: dict[str, tuple] = {
        "1": ("Fácil",   {"prob_turbo_rival": 0.10, "prob_obstaculo": 0.05,  "turbos_jugador": 4, "prob_turbo_bonus": 0.11}),
        "2": ("Normal",  {"prob_turbo_rival": 0.20, "prob_obstaculo": 0.15,  "turbos_jugador": 3, "prob_turbo_bonus": 0.08}),
        "3": ("Difícil", {"prob_turbo_rival": 0.35, "prob_obstaculo": 0.25,  "turbos_jugador": 2, "prob_turbo_bonus": 0.06}),
    }

    OBSTACULOS: list[tuple] = [
        ("⚠️  ¡Bache enorme! → retrocede 5m",     -5,  0),
        ("🛢️  ¡Derrame de aceite! → retrocede 10m", -10, 0),
        ("🚧  ¡Desvío obligatorio! → pierde 1 turbo", 0, -1),
        ("💨  ¡Viento a favor! → avanza 5m extra",  +5,  0),
        ("🌟  ¡Camino libre! → sin novedad",          0,  0),
    ]

    def __init__(self, distancia: int) -> None:
        self.distancia: int = distancia
        self.vehiculos: list[Vehiculo] = []
        self.turno_actual: int = 0
        self.dificultad_nombre: str = "Normal"
        self.cfg: dict = self.DIFICULTADES["2"][1]
        self.log: list[str] = []          # registro de eventos para la GUI

    @classmethod
    def crear_aleatoria(cls) -> "Carrera":
        """Crea una carrera con distancia aleatoria entre 100 y 2000m."""
        return cls(random.randint(100, 2000))

    @staticmethod
    def puntos_por_posicion(pos: int) -> int:
        """Devuelve los puntos según la posición al estilo F1. 0 si no puntúa."""
        return Carrera.PUNTOS_F1.get(pos, 0)

    def registrar_vehiculo(self, nombre_vehiculo: str, tipo: int) -> None:
        jugador = self._crear_vehiculo(nombre_vehiculo, tipo)
        jugador.es_jugador = True
        jugador.turbos = self.cfg["turbos_jugador"]
        self.vehiculos.append(jugador)

    def agregar_rivales_aleatorios(self) -> None:
        cantidad = random.randint(2, 6)
        contadores: dict[int, int] = {1: 0, 2: 0, 3: 0}
        nombres_tipo = {1: "Carro", 2: "Moto", 3: "Camión"}
        for _ in range(cantidad):
            tipo = random.choice([1, 2, 3])
            contadores[tipo] += 1
            nombre = f"{nombres_tipo[tipo]} {contadores[tipo]}"
            rival = self._crear_vehiculo(nombre, tipo)
            self.vehiculos.append(rival)

    def asignar_dificultad_aleatoria(self) -> None:
        clave = random.choice(["1", "2", "3"])
        self.dificultad_nombre, self.cfg = self.DIFICULTADES[clave]
        jugador = self._get_jugador()
        if jugador:
            jugador.turbos = self.cfg["turbos_jugador"]

    def ejecutar_turno_rivales(self) -> list[str]:
        """Mueve todos los rivales un turno. Retorna lista de eventos."""
        eventos: list[str] = []
        for v in self.vehiculos:
            if v.es_jugador:
                continue
            v.acelerar()
            if v.turbo_rival(self.cfg["prob_turbo_rival"]):
                eventos.append(f"🚀 {v.nombre} usó turbo! → {v.posicion}m")
            obs = self._aplicar_obstaculo(v)
            if obs:
                eventos.append(obs)
        self.turno_actual += 1
        return eventos

    def ejecutar_accion_jugador(self, usar_turbo: bool) -> list[str]:
        """Ejecuta la acción del jugador. Retorna lista de eventos."""
        jugador = self._get_jugador()
        if not jugador or jugador.posicion >= self.distancia:
            return []
        eventos: list[str] = []
        obs = self._aplicar_obstaculo(jugador)
        if obs:
            eventos.append(obs)
        antes = jugador.posicion
        jugador.acelerar()
        if usar_turbo and jugador.turbos > 0:
            jugador.usar_turbo()
            eventos.append(f"🚀 ¡TURBO! {antes}m → {jugador.posicion}m")
        else:
            eventos.append(f"✅ Avanzaste: {antes}m → {jugador.posicion}m")
        return eventos

    def _aplicar_obstaculo(self, vehiculo: Vehiculo) -> Optional[str]:
        etq = "¡Tú" if vehiculo.es_jugador else vehiculo.nombre
        mensajes: list[str] = []

        if random.random() < self.cfg["prob_turbo_bonus"]:
            vehiculo.turbos += 1
            mensajes.append(f"🎁 {etq}: ¡turbo extra encontrado! (turbos: {vehiculo.turbos})")

        if random.random() < self.cfg["prob_obstaculo"]:
            desc, delta_pos, delta_turbo = random.choice(self.OBSTACULOS)
            antes = vehiculo.posicion
            vehiculo.posicion = max(0, vehiculo.posicion + delta_pos)
            vehiculo.turbos = max(0, vehiculo.turbos + delta_turbo)
            icono = desc.split()[0]
            if delta_pos < 0:
                mensajes.append(f"{icono} {etq}: retrocede {abs(delta_pos)}m ({antes}m → {vehiculo.posicion}m)")
            elif delta_pos > 0:
                mensajes.append(f"{icono} {etq}: avanza {delta_pos}m extra ({antes}m → {vehiculo.posicion}m)")
            elif delta_turbo < 0:
                mensajes.append(f"{icono} {etq}: pierde 1 turbo (turbos: {vehiculo.turbos})")

        return "\n".join(mensajes) if mensajes else None

    def _get_jugador(self) -> Optional[Vehiculo]:
        for v in self.vehiculos:
            if v.es_jugador:
                return v
        return None

    def verificar_ganador(self) -> Optional[Vehiculo]:
        for v in sorted(self.vehiculos, key=lambda x: x.posicion, reverse=True):
            if v.posicion >= self.distancia:
                return v
        return None

    def ranking(self) -> list[Vehiculo]:
        return sorted(self.vehiculos, key=lambda v: v.posicion, reverse=True)

    def puntos_jugador(self) -> int:
        jugador = self._get_jugador()
        if not jugador:
            return 0
        ordenados = self.ranking()
        pos = ordenados.index(jugador) + 1
        return Carrera.puntos_por_posicion(pos)

    def posicion_jugador(self) -> int:
        jugador = self._get_jugador()
        if not jugador:
            return 0
        return self.ranking().index(jugador) + 1

    def obtener_datos_exportacion(self) -> list[dict]:
        """Devuelve la tabla de posiciones de esta carrera como lista de dicts."""
        ordenados = self.ranking()
        filas = []
        for i, v in enumerate(ordenados, 1):
            filas.append({
                "Posición":      i,
                "Nombre":        v.nombre,
                "Tipo":          type(v).__name__,
                "Distancia (m)": v.posicion,
                "Turbos usados": v.turbos_usados,
                "Puntos F1":     Carrera.puntos_por_posicion(i),
            })
        return filas

    def _crear_vehiculo(self, nombre: str, opcion: int) -> Vehiculo:
        if opcion == 1:
            return Carro(nombre, "Mazda", "ABC123", "Rojo", 120, 100)
        elif opcion == 2:
            return Moto(nombre, "Yamaha", "DEF456", "Azul", 150, 80)
        else:
            return Camion(nombre, "Chevrolet", "GHI789", "Verde", 90, 150)

    def __str__(self) -> str:
        return (f"Carrera(distancia={self.distancia}m, "
                f"participantes={len(self.vehiculos)}, "
                f"dificultad={self.dificultad_nombre})")

    def __repr__(self) -> str:
        return f"Carrera(distancia={self.distancia!r}, vehiculos={self.vehiculos!r})"

    def __len__(self) -> int:
        return len(self.vehiculos)


class Campeonato:
    """Gestiona N carreras consecutivas y el puntaje acumulado."""

    TOTAL_CARRERAS: int = 3

    def __init__(
        self,
        nombre_jugador: str,
        nombre_vehiculo: str,
        tipo_vehiculo: int,
    ) -> None:
        self.nombre_jugador: str = nombre_jugador
        self.nombre_vehiculo: str = nombre_vehiculo
        self.tipo_vehiculo: int = tipo_vehiculo
        self.carreras: list[Carrera] = []
        self.puntaje_total: int = 0
        self.resultados: list[dict] = []

    @classmethod
    def nuevo(cls, nombre: str, nombre_vehiculo: str, tipo: int) -> "Campeonato":
        """Constructor alternativo para crear un campeonato."""
        return cls(nombre, nombre_vehiculo, tipo)

    @staticmethod
    def formato_puntos(pts: int) -> str:
        """Formatea los puntos con signo explícito para mostrar en pantalla."""
        return f"+{pts}" if pts > 0 else str(pts)

    def registrar_resultado(self, carrera: Carrera) -> None:
        puntos = carrera.puntos_jugador()
        self.puntaje_total += puntos
        self.carreras.append(carrera)
        self.resultados.append({
            "carrera":    len(self.carreras),
            "distancia":  carrera.distancia,
            "dificultad": carrera.dificultad_nombre,
            "puntos":     puntos,
            "turnos":     carrera.turno_actual,
        })

    def clasificacion_acumulada(self) -> list[dict]:
        """Construye la tabla de clasificación final sumando datos de todas las carreras."""
        acumulado: dict[str, dict] = {}

        for carrera in self.carreras:
            for v in carrera.vehiculos:
                clave = v.nombre
                if clave not in acumulado:
                    acumulado[clave] = {
                        "Nombre":                v.nombre,
                        "Tipo":                  type(v).__name__,
                        "Carreras":              0,
                        "Distancia total (m)":   0,
                        "Puntos totales":        0,
                        "Turbos totales usados": 0,
                    }
                acumulado[clave]["Carreras"] += 1
                acumulado[clave]["Distancia total (m)"] += v.posicion

        for carrera in self.carreras:
            ordenados = sorted(carrera.vehiculos, key=lambda v: v.posicion, reverse=True)
            for i, v in enumerate(ordenados, 1):
                pts = Carrera.puntos_por_posicion(i)
                if v.nombre in acumulado:
                    acumulado[v.nombre]["Puntos totales"] += pts
                    acumulado[v.nombre]["Turbos totales usados"] += v.turbos_usados

        return sorted(
            acumulado.values(),
            key=lambda x: (x["Puntos totales"], x["Distancia total (m)"]),
            reverse=True,
        )

    def __str__(self) -> str:
        return (f"Campeonato(piloto={self.nombre_jugador!r}, "
                f"puntaje={self.puntaje_total})")

    def __repr__(self) -> str:
        return (f"Campeonato(nombre_jugador={self.nombre_jugador!r}, "
                f"puntaje_total={self.puntaje_total}, "
                f"carreras={len(self.carreras)})")


class Exportador:
    """
    Genera archivos de resultados con la información de todas las carreras.
    Soporta: CSV, JSON, TXT, XLSX (requiere openpyxl).
    """

    FORMATOS_DISPONIBLES: list[str] = ["csv", "json", "txt", "xlsx"]

    def __init__(self, campeonato: Campeonato) -> None:
        self.campeonato: Campeonato = campeonato

    def _nombre_archivo(self, formato: str) -> str:
        piloto = self.campeonato.nombre_jugador.lower().replace(" ", "_")
        return f"resultados_carreras_{piloto}.{formato}"

    def exportar(self, formato: str) -> str:
        """Exporta al formato indicado y retorna el nombre del archivo generado."""
        formato = formato.lower().strip()
        if formato == "csv":
            return self._exportar_csv()
        elif formato == "json":
            return self._exportar_json()
        elif formato == "txt":
            return self._exportar_txt()
        elif formato == "xlsx":
            return self._exportar_xlsx()
        else:
            raise ValueError(f"Formato '{formato}' no soportado.")

    def _exportar_csv(self) -> str:
        nombre = self._nombre_archivo("csv")
        with open(nombre, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            for i, carrera in enumerate(self.campeonato.carreras, 1):
                writer.writerow([f"=== CARRERA {i} | {carrera.distancia}m | {carrera.dificultad_nombre} ==="])
                filas = carrera.obtener_datos_exportacion()
                if filas:
                    writer.writerow(list(filas[0].keys()))
                    for fila in filas:
                        writer.writerow(list(fila.values()))
                writer.writerow([])

            writer.writerow(["=== RESUMEN DE CARRERAS (PILOTO) ==="])
            writer.writerow(["Carrera", "Distancia", "Dificultad", "Turnos", "Puntos F1"])
            for r in self.campeonato.resultados:
                writer.writerow([r["carrera"], f"{r['distancia']}m",
                                  r["dificultad"], r["turnos"], r["puntos"]])
            writer.writerow(["", "", "", "TOTAL", self.campeonato.puntaje_total])
            writer.writerow([])

            writer.writerow(["=== CLASIFICACIÓN ACUMULADA FINAL ==="])
            acumulada = self.campeonato.clasificacion_acumulada()
            if acumulada:
                writer.writerow(["Posición"] + list(acumulada[0].keys()))
                for pos, fila in enumerate(acumulada, 1):
                    writer.writerow([pos] + list(fila.values()))
        return nombre

    def _exportar_json(self) -> str:
        nombre = self._nombre_archivo("json")
        datos = {
            "piloto": self.campeonato.nombre_jugador,
            "vehiculo": self.campeonato.nombre_vehiculo,
            "puntaje_total": self.campeonato.puntaje_total,
            "carreras": [],
            "clasificacion_acumulada": self.campeonato.clasificacion_acumulada(),
        }
        for i, carrera in enumerate(self.campeonato.carreras, 1):
            datos["carreras"].append({
                "numero": i,
                "distancia": carrera.distancia,
                "dificultad": carrera.dificultad_nombre,
                "turnos": carrera.turno_actual,
                "puntos": self.campeonato.resultados[i - 1]["puntos"],
                "posiciones": carrera.obtener_datos_exportacion(),
            })
        with open(nombre, "w", encoding="utf-8") as f:
            json.dump(datos, f, ensure_ascii=False, indent=2)
        return nombre

    def _exportar_txt(self) -> str:
        ANCHO = 54
        nombre = self._nombre_archivo("txt")
        lineas: list[str] = []
        lineas.append("=" * ANCHO)
        lineas.append("RESULTADOS DEL CAMPEONATO".center(ANCHO))
        lineas.append(f"Piloto: {self.campeonato.nombre_jugador}".center(ANCHO))
        lineas.append("=" * ANCHO)

        for i, carrera in enumerate(self.campeonato.carreras, 1):
            lineas.append(f"\n--- CARRERA {i} | {carrera.distancia}m | {carrera.dificultad_nombre} ---")
            filas = carrera.obtener_datos_exportacion()
            if filas:
                enc = list(filas[0].keys())
                lineas.append("  " + " | ".join(f"{k:<15}" for k in enc))
                lineas.append("  " + "-" * (17 * len(enc)))
                for fila in filas:
                    lineas.append("  " + " | ".join(f"{str(v):<15}" for v in fila.values()))

        lineas.append("\n" + "=" * ANCHO)
        lineas.append("RESUMEN DE CARRERAS (PILOTO)".center(ANCHO))
        lineas.append("=" * ANCHO)
        for r in self.campeonato.resultados:
            lineas.append(
                f"  Carrera {r['carrera']}: {r['distancia']}m — "
                f"{r['dificultad']} — {r['turnos']} turnos → {r['puntos']} pts"
            )
        lineas.append(f"\n  PUNTAJE TOTAL: {self.campeonato.puntaje_total} pts")

        lineas.append("\n" + "=" * ANCHO)
        lineas.append("CLASIFICACIÓN ACUMULADA FINAL".center(ANCHO))
        lineas.append("=" * ANCHO)
        acumulada = self.campeonato.clasificacion_acumulada()
        medallas = {1: "1°", 2: "2°", 3: "3°"}
        enc = ["Pos", "Nombre", "Tipo", "Dist.total", "Pts F1", "Turbos"]
        lineas.append("  " + " | ".join(f"{e:<13}" for e in enc))
        lineas.append("  " + "-" * (15 * len(enc)))
        for pos, fila in enumerate(acumulada, 1):
            med = medallas.get(pos, f"{pos}°")
            lineas.append(
                f"  {str(pos) + ' ' + med:<13} "
                f"| {fila['Nombre']:<13} "
                f"| {fila['Tipo']:<13} "
                f"| {str(fila['Distancia total (m)']) + 'm':<13} "
                f"| {fila['Puntos totales']:<13} "
                f"| {fila['Turbos totales usados']:<13}"
            )
        lineas.append("=" * ANCHO)

        with open(nombre, "w", encoding="utf-8") as f:
            f.write("\n".join(lineas))
        return nombre

    def _exportar_xlsx(self) -> str:
        if not OPENPYXL_DISPONIBLE:
            raise RuntimeError("openpyxl no está instalado. Ejecuta: pip install openpyxl")
        nombre = self._nombre_archivo("xlsx")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        estilo_enc = Font(bold=True, color="FFFFFF")
        fondo_enc  = PatternFill("solid", start_color="1F4E79")
        centro     = Alignment(horizontal="center")

        def escribir_tabla(ws, filas: list[dict], fila_inicio: int = 1) -> int:
            if not filas:
                return fila_inicio
            encabezados = list(filas[0].keys())
            for col, enc in enumerate(encabezados, 1):
                c = ws.cell(row=fila_inicio, column=col, value=enc)
                c.font = estilo_enc
                c.fill = fondo_enc
                c.alignment = centro
            for fila_idx, fila in enumerate(filas, fila_inicio + 1):
                for col, valor in enumerate(fila.values(), 1):
                    ws.cell(row=fila_idx, column=col, value=valor)
            return fila_inicio + len(filas) + 2

        for i, carrera in enumerate(self.campeonato.carreras, 1):
            ws = wb.create_sheet(title=f"Carrera {i}")
            ws["A1"] = f"Carrera {i} — {carrera.distancia}m — {carrera.dificultad_nombre}"
            ws["A1"].font = Font(bold=True, size=13)
            for col_letra, ancho in zip(["A","B","C","D","E","F"], [6, 18, 12, 14, 14, 12]):
                ws.column_dimensions[col_letra].width = ancho
            filas = carrera.obtener_datos_exportacion()
            escribir_tabla(ws, filas, fila_inicio=3)

        ws_f = wb.create_sheet(title="Clasificación Final")
        ws_f["A1"] = f"Campeonato — Piloto: {self.campeonato.nombre_jugador}"
        ws_f["A1"].font = Font(bold=True, size=13)
        ws_f["A3"] = "Resumen de carreras (piloto)"
        ws_f["A3"].font = Font(bold=True)
        resumen_filas = [
            {
                "Carrera":    r["carrera"],
                "Distancia":  f"{r['distancia']}m",
                "Dificultad": r["dificultad"],
                "Turnos":     r["turnos"],
                "Puntos F1":  r["puntos"],
            }
            for r in self.campeonato.resultados
        ]
        sig_fila = escribir_tabla(ws_f, resumen_filas, fila_inicio=4)
        ws_f.cell(row=sig_fila, column=4, value="TOTAL").font = Font(bold=True)
        ws_f.cell(row=sig_fila, column=5, value=self.campeonato.puntaje_total).font = Font(bold=True)
        sig_fila += 2

        ws_f.cell(row=sig_fila, column=1, value="Clasificación acumulada final").font = Font(bold=True)
        sig_fila += 1
        acumulada = self.campeonato.clasificacion_acumulada()
        acumulada_pos = [{"Posición": i, **fila} for i, fila in enumerate(acumulada, 1)]
        escribir_tabla(ws_f, acumulada_pos, fila_inicio=sig_fila)

        for col in ["A","B","C","D","E","F","G"]:
            ws_f.column_dimensions[col].width = 22

        wb.save(nombre)
        return nombre

    def __str__(self) -> str:
        return f"Exportador(piloto={self.campeonato.nombre_jugador!r})"

    def __repr__(self) -> str:
        return f"Exportador(campeonato={self.campeonato!r})"
